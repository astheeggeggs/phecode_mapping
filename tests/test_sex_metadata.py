from __future__ import annotations

import csv
from pathlib import Path

import pytest
from openpyxl import Workbook

from conftest import write_csv
from phecodex_mapper.sex_metadata import enrich_sex_metadata


def test_enrichment_assigns_direct_moved_new_and_semantic_values(tmp_path: Path) -> None:
    old = tmp_path / "old.csv"; new = tmp_path / "new.csv"; changes = tmp_path / "changes.xlsx"
    write_csv(old, ["phecode", "phecode_string", "sex"], [["AA_1", "old", "Female"], ["OLD_1", "old", "Male"]])
    write_csv(new, ["phecode", "phecode_string"], [["AA_1", "same"], ["NEW_1", "new"], ["PP_P001", "pregnancy"]])
    book = Workbook(); book.remove(book.active)
    moved = book.create_sheet("Phecodes moved"); moved.append(["old", "old string", "new", "new string", "reason"]); moved.append(["OLD_1", "old", "NEW_1", "new", "move"])
    added = book.create_sheet("New phecodes"); added.append(["phecode", "phecode_string", "Category", "sex", "icd10_only"])
    book.save(changes)
    output = tmp_path / "out.csv"; review = tmp_path / "review.csv"
    counts = enrich_sex_metadata(old, new, changes, output, review)
    with output.open(newline="") as stream: rows = {r["phecode"]: r for r in csv.DictReader(stream)}
    assert rows["AA_1"]["sex"] == "Female"
    assert rows["NEW_1"]["sex"] == "Male"
    assert rows["PP_P001"]["sex"] == "Female"
    assert counts == {"Both": 0, "Female": 2, "Male": 1}
    assert "semantic: pregnancy phenotype" in review.read_text()


def test_enrichment_rejects_unassigned_phecodes(tmp_path: Path) -> None:
    old = tmp_path / "old.csv"; new = tmp_path / "new.csv"; changes = tmp_path / "changes.xlsx"
    write_csv(old, ["phecode", "sex"], [["AA_1", "Both"]]); write_csv(new, ["phecode"], [["UNKNOWN", ""]])
    book = Workbook(); book.active.title = "Phecodes moved"; book.create_sheet("New phecodes"); book.save(changes)
    with pytest.raises(ValueError, match="No sex assignment"):
        enrich_sex_metadata(old, new, changes, tmp_path / "out.csv")
