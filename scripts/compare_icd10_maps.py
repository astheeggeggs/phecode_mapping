#!/usr/bin/env python3
from __future__ import annotations
import argparse, csv, re, json
from collections import defaultdict, Counter
from pathlib import Path
from openpyxl import load_workbook

def load_map(path: Path, code_column: str, vocabulary: str) -> dict[str, set[str]]:
    result = defaultdict(set)
    with path.open(newline="", encoding="latin1") as stream:
        for row in csv.DictReader(stream):
            if row["vocabulary_id"].upper() == vocabulary:
                code = re.sub(r"[.\s-]", "", row[code_column].strip().upper())
                result[code].add(row["phecode"])
    return result

parser = argparse.ArgumentParser(description="Compare PhecodeX WHO and ICD-10-CM maps.")
parser.add_argument("--cm-map", type=Path, required=True); parser.add_argument("--who-map", type=Path, required=True)
parser.add_argument("--changes", type=Path, required=True); parser.add_argument("--output", type=Path, required=True)
parser.add_argument("--summary", type=Path, required=True); parser.add_argument("--shared-output", type=Path)
parser.add_argument("--shared-summary", type=Path); args = parser.parse_args()
cm = load_map(args.cm_map, "icd", "ICD10CM"); who = load_map(args.who_map, "ICD", "ICD10")
wb = load_workbook(args.changes, read_only=True, data_only=True)
moved = {r[0]: r[2] for r in wb["Phecodes moved"].iter_rows(min_row=2, values_only=True)}
new_phecodes = {r[0] for r in wb["New phecodes"].iter_rows(min_row=2, values_only=True)}
added = {re.sub(r"[.\s-]", "", str(r[0]).strip().upper()) for r in wb["ICDs added"].iter_rows(min_row=2, values_only=True)}
removed = {re.sub(r"[.\s-]", "", str(r[0]).strip().upper()) for r in wb["ICDs remoed"].iter_rows(min_row=2, values_only=True)}
rows=[]; counts=Counter()
for code in sorted(set(cm)|set(who)):
    c, w = cm.get(code,set()), who.get(code,set())
    if c == w: continue
    if not c: kind="WHO-only code"
    elif not w: kind="CM-only code"
    else: kind="shared code: mapping differs"
    reasons=[]
    if code in added: reasons.append("ICD listed as added")
    if code in removed: reasons.append("ICD listed as removed")
    if c and w and any(moved.get(old) in (c-w) for old in (w-c)): reasons.append("phecode moved")
    if (c-w) and (c-w) <= new_phecodes: reasons.append("new phecode")
    counts[kind] += 1
    rows.append({"normalized_icd10":code,"difference_type":kind,"icd10cm_phecodes":"|".join(sorted(c)),"who_icd10_phecodes":"|".join(sorted(w)),"possible_change_explanation":"; ".join(reasons)})
args.output.parent.mkdir(parents=True, exist_ok=True)
with args.output.open("w",newline="",encoding="utf-8") as stream:
    writer=csv.DictWriter(stream,fieldnames=list(rows[0])); writer.writeheader(); writer.writerows(rows)
args.summary.write_text(json.dumps({"rows":len(rows),"difference_types":counts,"shared_codes":len(set(cm)&set(who)),"cm_unique_codes":len(cm),"who_unique_codes":len(who)},indent=2,sort_keys=True,default=dict)+"\n")
shared = [r for r in rows if r["difference_type"] == "shared code: mapping differs"]
if args.shared_output:
    args.shared_output.parent.mkdir(parents=True, exist_ok=True)
    with args.shared_output.open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=list(shared[0])); writer.writeheader(); writer.writerows(shared)
if args.shared_summary:
    signatures = Counter((r["icd10cm_phecodes"], r["who_icd10_phecodes"]) for r in shared)
    args.shared_summary.write_text(json.dumps({"rows": len(shared), "signature_count": len(signatures), "signatures": [{"icd10cm_phecodes": c, "who_icd10_phecodes": w, "code_count": n} for (c, w), n in signatures.most_common()]}, indent=2)+"\n")
print(f"rows: {len(rows)}\nsummary: {args.summary}")
