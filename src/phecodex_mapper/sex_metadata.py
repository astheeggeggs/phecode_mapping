from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

VALID_SEX = {"Both", "Male", "Female"}
SEMANTIC_ASSIGNMENTS = {
    "BI_161.1": ("Both", "semantic: hematologic phenotype"), "BI_161.3": ("Both", "semantic: hematologic phenotype"), "BI_161.4": ("Both", "semantic: hematologic phenotype"),
    "NB_N000": ("Both", "semantic: neonatal phenotype"), "NB_N000.1": ("Both", "semantic: neonatal phenotype"), "NB_N000.2": ("Both", "semantic: neonatal phenotype"), "NB_N000.3": ("Both", "semantic: neonatal phenotype"), "NB_N000.4": ("Both", "semantic: neonatal phenotype"), "NB_N000.5": ("Both", "semantic: neonatal phenotype"), "NB_N000.6": ("Both", "semantic: neonatal phenotype"), "NB_N000.7": ("Both", "semantic: neonatal phenotype"), "NB_N000.71": ("Both", "semantic: neonatal phenotype"), "NB_N000.72": ("Both", "semantic: neonatal phenotype"), "NB_N010": ("Both", "semantic: neonatal phenotype"),
    "PP_P001": ("Female", "semantic: pregnancy phenotype"), "PP_P002": ("Female", "semantic: pregnancy phenotype"), "PP_P003": ("Female", "semantic: pregnancy phenotype"),
}

def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="latin1") as stream:
        return list(csv.DictReader(stream))

def enrich_sex_metadata(v1_info: Path, v11_info: Path, changes: Path, output: Path, review_output: Path | None = None) -> dict[str, int]:
    old, new = _read_csv(v1_info), _read_csv(v11_info)
    old_by_code, new_by_code = {r["phecode"]: r for r in old}, {r["phecode"]: r for r in new}
    if len(old_by_code) != len(old) or len(new_by_code) != len(new): raise ValueError("Metadata contains duplicate phecode rows")
    if not old or "sex" not in old[0]: raise ValueError("1.0 metadata must contain a sex column")
    if not new or "sex" in new[0]: raise ValueError("1.1 metadata should not already contain a sex column")
    assignments: dict[str, tuple[str, str]] = {c: (old_by_code[c]["sex"].strip().title(), "direct 1.0 match") for c in new_by_code if c in old_by_code}
    workbook = load_workbook(changes, read_only=True, data_only=True)
    for row in workbook["Phecodes moved"].iter_rows(min_row=2, values_only=True):
        source, target = row[0], row[2]
        if target in new_by_code and source in old_by_code: assignments[target] = (old_by_code[source]["sex"].strip().title(), f"moved from {source}")
    for row in workbook["New phecodes"].iter_rows(min_row=2, values_only=True):
        code, sex = row[0], row[3]
        if code in new_by_code: assignments[code] = (str(sex).strip().title(), "workbook New phecodes")
    for code, value in SEMANTIC_ASSIGNMENTS.items():
        if code in new_by_code: assignments[code] = value
    missing = set(new_by_code) - set(assignments)
    if missing: raise ValueError(f"No sex assignment for phecodes: {sorted(missing)}")
    invalid = {c: v for c, v in assignments.items() if v[0] not in VALID_SEX}
    if invalid: raise ValueError(f"Invalid sex assignments: {invalid}")
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = list(new[0]) + ["sex"]
    with output.open("w", newline="", encoding="utf-8") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields); writer.writeheader()
        for row in new:
            result = dict(row); result["sex"] = assignments[row["phecode"]][0]; writer.writerow(result)
    review = [{"phecode": c, "sex": assignments[c][0], "source": assignments[c][1], "phecode_string": new_by_code[c].get("phecode_string", ""), "category": new_by_code[c].get("category", "")} for c in sorted(new_by_code) if c not in old_by_code or assignments[c][1] != "direct 1.0 match"]
    if review_output:
        review_output.parent.mkdir(parents=True, exist_ok=True)
        with review_output.open("w", newline="", encoding="utf-8") as stream:
            writer = csv.DictWriter(stream, fieldnames=list(review[0])); writer.writeheader(); writer.writerows(review)
    return {sex: sum(value[0] == sex for value in assignments.values()) for sex in sorted(VALID_SEX)}
